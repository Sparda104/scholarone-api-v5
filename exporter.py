# exporter.py - V5 ENHANCED with Data Pipeline
"""
Enhanced Excel Exporter with V5 Data Pipeline:
- JSON flattening with dot notation (Standing Order #9)
- Array explosion to separate rows (Standing Order #9)
- Professional formatting and comprehensive JSON handling
"""

from __future__ import annotations
import os
import re
import json
from typing import Any, Dict, Iterable, List, Optional, Sequence, Union
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import sanitization function from utils
try:
    from utils import sanitize_filename as util_sanitize_filename
    HAS_UTIL_SANITIZE = True
except ImportError:
    HAS_UTIL_SANITIZE = False

JsonRow = Dict[str, Any]
RowsLike = Iterable[Union[JsonRow, Any]]


# ==================== V5 DATA PIPELINE FUNCTIONS ====================

def _flatten_dict(data: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
    """
    Flatten nested dictionaries with dot notation.
    Standing Order #9: Convert nested JSON to flat structure.

    Example:
        {"author": {"firstName": "John"}} -> {"author.firstName": "John"}
    """
    items = []
    for k, v in data.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            # Recursively flatten nested dicts
            items.extend(_flatten_dict(v, new_key, sep).items())
        elif isinstance(v, list):
            # Preserve arrays for explosion
            items.append((new_key, v))
        else:
            items.append((new_key, v))
    return dict(items)


def _explode_arrays(rows: List[JsonRow]) -> List[JsonRow]:
    """
    Explode arrays to separate rows for relational data.
    Standing Order #9: If submission has 3 authors, create 3 rows.

    Example:
        {"id": "1", "authors": [{"name": "A1"}, {"name": "A2"}]}
        ->
        [{"id": "1", "authors.name": "A1"},
         {"id": "1", "authors.name": "A2"}]
    """
    exploded_rows = []

    for row in rows:
        # Find array fields
        array_fields = {k: v for k, v in row.items() 
                       if isinstance(v, list) and v}

        if not array_fields:
            # No arrays, keep as is
            exploded_rows.append(row)
            continue

        # Get max array length
        max_len = max(len(v) for v in array_fields.values())

        # Create separate row for each array index
        for i in range(max_len):
            new_row = {}
            for k, v in row.items():
                if isinstance(v, list) and v:
                    # Array field
                    idx = i if i < len(v) else len(v) - 1
                    if isinstance(v[idx], dict):
                        # Flatten dict elements in array
                        for sub_k, sub_v in v[idx].items():
                            new_row[f"{k}.{sub_k}"] = sub_v
                    else:
                        new_row[k] = v[idx]
                else:
                    # Non-array field, repeat in each row
                    new_row[k] = v
            exploded_rows.append(new_row)

    return exploded_rows


def _detect_json_malfunction(rows: List[JsonRow]) -> None:
    """
    Detect remaining JSON strings in cells (indicates pipeline failure).
    Logs warnings if JSON strings found.
    """
    for i, row in enumerate(rows[:5]):  # Check first 5 rows
        for key, value in row.items():
            if isinstance(value, str):
                stripped = value.strip()
                if ((stripped.startswith('{') and stripped.endswith('}')) or
                    (stripped.startswith('[') and stripped.endswith(']'))):
                    print(f"[WARN] JSON string detected in row {i}, field '{key}'")
                    break

# =====================================================================


def _is_primitive(x: Any) -> bool:
    return isinstance(x, (str, int, float, bool)) or x is None


def _cellify(val: Any) -> Any:
    """Convert any Python/JSON value into something Excel can store."""
    if _is_primitive(val):
        return val

    # Handle common nested structures
    if isinstance(val, dict):
        if "name" in val:
            return val["name"]
        elif "id" in val:
            return f"ID: {val['id']}"
        elif len(val) == 1:
            key, value = next(iter(val.items()))
            return f"{key}: {value}"

    try:
        return json.dumps(val, ensure_ascii=False)
    except Exception:
        return str(val)


def _coerce_row_to_dict(row: Any) -> JsonRow:
    """Guarantee a dict row. Wrap scalars/unknowns under 'submission' key."""
    if isinstance(row, dict):
        return row

    # Try to parse JSON strings
    if isinstance(row, str):
        s = row.strip()
        if (s.startswith("{") and s.endswith("}")) or (s.startswith("[") and s.endswith("]")):
            try:
                parsed = json.loads(s)
                if isinstance(parsed, dict):
                    return parsed
                return {"submission": parsed}
            except Exception:
                pass

    return {"submission": row}


def _infer_columns(rows: Iterable[JsonRow]) -> List[str]:
    """Collect all keys across rows, preserving first-seen order."""
    seen = {}
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen[k] = True

    keys = list(seen.keys()) or ["submission"]

    # Ensure Journal column is first if it exists
    if 'Journal' in keys:
        keys.remove('Journal')
        keys.insert(0, 'Journal')

    return keys


def _sanitize_filename(name: str) -> str:
    """Sanitize filename for safe filesystem usage."""
    if HAS_UTIL_SANITIZE:
        try:
            sanitized = util_sanitize_filename(name)
            if not sanitized.lower().endswith('.xlsx'):
                sanitized += '.xlsx'
            return sanitized
        except Exception:
            pass

    # Fallback: Local sanitization
    invalid = set(r'<>:"/\|?*')
    safe = "".join(("_" if ch in invalid else ch) for ch in name)
    safe = re.sub(r'\s+', ' ', safe).strip()
    return safe or "export.xlsx"


def _auto_adjust_column_width(worksheet, max_width=50):
    """Auto-adjust column widths based on content."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 10), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


class ExcelExporter:
    def __init__(self, logger=None):
        self.logger = logger

    def flatten(self, rows: RowsLike) -> List[JsonRow]:
        """Flatten/normalize rows for testing."""
        if not rows:
            return []
        return [_coerce_row_to_dict(r) for r in rows]

    def export_to_excel(
        self,
        rows: RowsLike,
        filename: str,
        columns: Optional[Sequence[str]] = None,
        export_dir: Optional[str] = None,
        apply_formatting: bool = True,
        enable_pipeline: bool = True  # V5: Enable data pipeline
    ) -> str:
        """
        Write all rows to a single-sheet Excel file with V5 data pipeline.

        Args:
            rows: Iterable of dict-like JSON rows
            filename: desired file name (will be sanitized)
            columns: optional explicit column order; if None, inferred
            export_dir: optional directory; if None, uses default
            apply_formatting: whether to apply Excel formatting
            enable_pipeline: whether to apply V5 data pipeline (flatten + explode)

        Returns:
            str: Full path to the created Excel file
        """
        # Determine export directory
        if export_dir is None:
            export_dir = os.environ.get(
                'SCHOLARONE_EXPORT_DIR',
                os.path.join(os.path.expanduser("~"), "ScholarOne_Exports")
            )

        out_dir = export_dir
        if not isinstance(out_dir, str):
            raise TypeError("export_dir must be a string path or None")

        os.makedirs(out_dir, exist_ok=True)

        # Sanitize filename
        base = filename if filename.lower().endswith(".xlsx") else f"{filename}.xlsx"
        base = _sanitize_filename(base)
        out_path = os.path.join(out_dir, base)

        # Coerce rows to dictionaries
        normalized: List[JsonRow] = [_coerce_row_to_dict(r) for r in rows]

        if not normalized:
            # Create empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(["No data found"])
            wb.save(out_path)
            return out_path

        # ========== V5 DATA PIPELINE ==========
        if enable_pipeline:
            if self.logger:
                self.logger.info("Applying V5 data pipeline (flatten + explode)")

            # Step 1: Flatten nested dictionaries
            normalized = [_flatten_dict(row) for row in normalized]

            # Step 2: Explode arrays to separate rows
            normalized = _explode_arrays(normalized)

            # Step 3: Detect any remaining JSON (malfunction check)
            _detect_json_malfunction(normalized)
        # ======================================

        # Determine column order
        cols: List[str] = list(columns) if columns else _infer_columns(normalized)

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "ScholarOne Export"

        # Header row
        ws.append(cols)

        if apply_formatting:
            # Format header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_num, _ in enumerate(cols, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        # Data rows
        for row_num, r in enumerate(normalized, start=2):
            row_data = []
            for col in cols:
                value = r.get(col, "")
                cellified_value = _cellify(value)
                row_data.append(cellified_value)
            ws.append(row_data)

            # Zebra striping
            if apply_formatting and row_num % 2 == 0:
                for col_num in range(1, len(cols) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        if apply_formatting:
            # Borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                   min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # Auto-adjust column widths
            _auto_adjust_column_width(ws)

            # Freeze header row
            ws.freeze_panes = "A2"

        # Add summary sheet for large exports
        if len(normalized) > 10:
            summary_ws = wb.create_sheet("Summary")
            summary_ws.append(["Export Summary"])
            summary_ws.append(["Total Records", len(normalized)])
            summary_ws.append(["Total Columns", len(cols)])
            summary_ws.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            summary_ws.append(["Pipeline Enabled", "Yes" if enable_pipeline else "No"])
            summary_ws.append([])
            summary_ws.append(["Column Names"])
            for col in cols:
                summary_ws.append([col])

            if apply_formatting:
                summary_ws.cell(1, 1).font = Font(bold=True, size=14)
                for row in range(2, 6):
                    summary_ws.cell(row, 1).font = Font(bold=True)

        wb.save(out_path)

        if self.logger:
            try:
                pipeline_note = " (with pipeline)" if enable_pipeline else ""
                self.logger.info(f"Exported {len(normalized)} rows to {out_path}{pipeline_note}")
            except Exception:
                pass

        return out_path

    def export_multiple_sheets(
        self,
        data_dict: Dict[str, RowsLike],
        filename: str,
        export_dir: Optional[str] = None
    ) -> str:
        """
        Export multiple datasets to separate sheets in one Excel file.

        Args:
            data_dict: Dictionary where keys are sheet names and values are row data
            filename: desired file name
            export_dir: optional directory; if None, uses default

        Returns:
            str: Full path to the created Excel file
        """
        # Determine export directory
        if export_dir is None:
            export_dir = os.environ.get(
                'SCHOLARONE_EXPORT_DIR',
                os.path.join(os.path.expanduser("~"), "ScholarOne_Exports")
            )

        out_dir = export_dir
        os.makedirs(out_dir, exist_ok=True)

        base = filename if filename.lower().endswith(".xlsx") else f"{filename}.xlsx"
        base = _sanitize_filename(base)
        out_path = os.path.join(out_dir, base)

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for sheet_name, rows in data_dict.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
            normalized: List[JsonRow] = [_coerce_row_to_dict(r) for r in rows]

            if not normalized:
                ws.append(["No data found"])
                continue

            cols = _infer_columns(normalized)
            ws.append(cols)

            for r in normalized:
                ws.append([_cellify(r.get(c, "")) for c in cols])

            _auto_adjust_column_width(ws)

        wb.save(out_path)

        if self.logger:
            try:
                self.logger.info(f"Exported {len(data_dict)} sheets to {out_path}")
            except Exception:
                pass

        return out_path
